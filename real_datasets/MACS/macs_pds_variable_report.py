import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

with open('/tmp/macs_report.json') as f:
    report = json.load(f)

wb = Workbook()

# Colors
HEADER_FILL = PatternFill('solid', fgColor='13294B')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
CAT_FILL = PatternFill('solid', fgColor='FF5F05')
CAT_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
DERIVED_FONT = Font(name='Arial', size=10, italic=True, color='666666')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC')
)

# ============ Sheet 1: Full Variable Inventory ============
ws = wb.active
ws.title = "Variable Inventory"

headers = ['SILK Category', 'Variable', 'Source File', 'Description', 'SILK Role',
           'N (obs)', 'Mean', 'SD', 'Median', 'Q25', 'Q75', 'Min', 'Max', '% Non-missing', 'Notes']

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:O1"

cat_labels = {
    '1_ID': '1. Subject ID',
    '2_TIME_FROM_ONSET': '2. Estimated Time-from-Onset',
    '3_BIOMARKER_AGE': '3. Biomarker Measurement Age',
    '4_BIOMARKERS': '4. Longitudinal Biomarkers',
    '5_EVENT': '5. Event / Survival Endpoint',
    '6_COVARIATES': '6. Non-Time-Varying Covariates',
}

row = 2
for cat_key, items in report.items():
    cat_label = cat_labels.get(cat_key, cat_key)
    
    # Category separator row
    ws.cell(row=row, column=1, value=cat_label).font = CAT_FONT
    ws.cell(row=row, column=1).fill = CAT_FILL
    for c in range(2, len(headers)+1):
        ws.cell(row=row, column=c).fill = CAT_FILL
    row += 1
    
    for item in items:
        is_derived = item.get('file','') == 'derived'
        font = DERIVED_FONT if is_derived else BODY_FONT
        
        ws.cell(row=row, column=1, value=cat_label.split('. ',1)[-1]).font = font
        ws.cell(row=row, column=2, value=item.get('variable','')).font = BOLD_FONT if not is_derived else DERIVED_FONT
        ws.cell(row=row, column=3, value=item.get('file','')).font = font
        ws.cell(row=row, column=4, value=item.get('description','')).font = font
        ws.cell(row=row, column=5, value=item.get('silk_role','')).font = font
        
        # Stats
        n = item.get('n', item.get('n_converters', item.get('n_total', '')))
        ws.cell(row=row, column=6, value=n if n != '' else None).font = font
        ws.cell(row=row, column=7, value=item.get('mean')).font = font
        ws.cell(row=row, column=8, value=item.get('sd')).font = font
        ws.cell(row=row, column=9, value=item.get('median')).font = font
        ws.cell(row=row, column=10, value=item.get('q25')).font = font
        ws.cell(row=row, column=11, value=item.get('q75')).font = font
        ws.cell(row=row, column=12, value=item.get('min')).font = font
        ws.cell(row=row, column=13, value=item.get('max')).font = font
        ws.cell(row=row, column=14, value=item.get('pct_nonmissing')).font = font
        
        # Notes
        notes = []
        if is_derived: notes.append('Derived variable')
        if 'distribution' in item:
            dist_str = '; '.join(f"{k}: {v}" for k,v in item['distribution'].items())
            notes.append(dist_str)
        if 'n_events' in item:
            notes.append(f"Events: {item['n_events']}/{item.get('n_total','?')} ({item.get('event_rate_pct','?')}%)")
        if 'note' in item:
            notes.append(item['note'])
        if item.get('n_converters'):
            notes.append(f"N converters={item['n_converters']}, N total MACS={item.get('n_total_macs','')}")
        ws.cell(row=row, column=15, value=' | '.join(notes) if notes else None).font = font
        
        for c in range(1, len(headers)+1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

# Column widths
col_widths = [28, 16, 22, 55, 32, 10, 10, 10, 10, 10, 10, 10, 10, 12, 65]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ============ Sheet 2: SILK Data Mapping ============
ws2 = wb.create_sheet("SILK Mapping")

mapping = [
    ['SILK Notation', 'SILK Concept', 'MACS PDS Variable(s)', 'Construction', 'Notes'],
    ['i', 'Subject index', 'CASEID', 'Direct', 'N=603 seroconverters'],
    ['ε_i', 'Common origin shift', 'SC_INTERVAL / 2', 'True infection ∈ [V1DATY, V2DATY]; ε_i = true − midpoint', 'Bounded by |ε_i| ≤ SC_INTERVAL/2; median interval = 1 yr → |ε_i| ≤ 0.5 yr'],
    ['A*_i', 'True onset age', 'Unknown', 'True HIV infection date — unobserved', 'Latent; this IS the origin error'],
    ['Â_i = A*_i + ε_i', 'Estimated onset date', 'SC_MIDPOINT = (V1DATY + V2DATY) / 2', 'Midpoint of seroconversion interval', 'Standard convention in HIV literature'],
    ['A_ij', 'Observed biomarker age', 'TIME_SINCE_SC = LDATY − SC_MIDPOINT', 'Calendar year of visit minus estimated SC year', 'Shifted by ε_i from true disease age'],
    ['W_ij', 'Biomarker vector at visit j', 'LEU3N, LEU3P, LEU2N, LEU2P, VLOAD, ...', 'lab_rslt.dat columns at each visit', 'Primary: CD4, CD8, viral load; Secondary: CBC, lipids, liver'],
    ['T_i', 'Event time', 'DATE1yy − SC_MIDPOINT', 'Years from estimated SC to AIDS diagnosis', 'Among events: mean ~3 yr post-SC'],
    ['δ_i', 'Event indicator', 'AIDSCASE ∈ {2,3}', '1 if AIDS diagnosed, 0 otherwise', '292/603 = 48.4% event rate'],
    ['C_i', 'Censoring time', 'max(LDATY) − SC_MIDPOINT', 'Years from estimated SC to last visit (if no AIDS)', 'Median follow-up 8.5 yr'],
    ['W_i', 'Baseline covariates', 'AGE_AT_SC, RACE, EDUCA, MACSCODE', 'From macsid.dat and section2.dat baseline', 'Non-time-varying'],
]

for r, row_data in enumerate(mapping, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        else:
            cell.font = BOLD_FONT if c <= 2 else BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = THIN_BORDER

for i, w in enumerate([16, 24, 36, 50, 55], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A2'

# ============ Sheet 3: Biomarker Recommendation ============
ws3 = wb.create_sheet("Biomarker Selection")

recs = [
    ['Variable', 'Type', '% Available', 'Median per Subj', 'Recommendation', 'Rationale'],
    ['LEU3N (CD4 count)', 'Primary', '95.7%', '20', 'USE — core', 'Gold-standard marker of immunosuppression; monotone decline with disease progression; directly tied to AIDS definition (<200)'],
    ['LEU3P (CD4 %)', 'Primary', '96.5%', '~20', 'USE — core', 'Proportion-based; more stable than count; also in AIDS definition (<14%)'],
    ['LEU2N (CD8 count)', 'Primary', '95.7%', '~20', 'USE — core', 'Rises with HIV viremia, tracks immune activation; CD4/CD8 ratio is a key prognostic index'],
    ['VLOAD (viral load)', 'Primary', '66.7%', '~11', 'USE — core', 'Direct measure of viral replication; strong predictor of progression; 33% missing (pre-assay era)'],
    ['PLATE (platelets)', 'Secondary', '98.4%', '~20', 'USE — enrichment', 'Thrombocytopenia is HIV-associated; high availability; adds distributional variety'],
    ['HCT (hematocrit)', 'Secondary', '99.1%', '~21', 'USE — enrichment', 'Anemia tracks disease severity; near-complete coverage'],
    ['WBC', 'Secondary', '99.2%', '~21', 'CONSIDER', 'Leukopenia in advanced HIV; available but less specific'],
    ['TCHOL (cholesterol)', 'Secondary', '42.2%', '~10', 'OPTIONAL', 'Metabolic marker; >50% missing limits utility; available mainly post-1996'],
    ['SGPT/SGOT (ALT/AST)', 'Secondary', '38.4%', '~10', 'OPTIONAL', 'Liver function; relevant for hepatitis co-infection subgroup'],
    ['CREAT (creatinine)', 'Secondary', '30.9%', '~8', 'SKIP', 'Too much missing data for reliable longitudinal modeling'],
    ['GLUC2, HGA1C', 'Secondary', '30-40%', '~8', 'SKIP', 'Metabolic; high missingness; not directly HIV-stage-informative'],
]

for r, row_data in enumerate(recs, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        else:
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if c == 5:
                if 'core' in val: cell.fill = PatternFill('solid', fgColor='C6EFCE')
                elif 'enrichment' in val: cell.fill = PatternFill('solid', fgColor='FFEB9C')
                elif 'SKIP' in val: cell.fill = PatternFill('solid', fgColor='FFC7CE')
        cell.border = THIN_BORDER

for i, w in enumerate([22, 12, 12, 14, 18, 65], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = 'A2'

out = "/sessions/lucid-brave-newton/mnt/SILK/real_datasets/MACS_PDS_Variable_Report.xlsx"
wb.save(out)
print(f"Saved: {out}")
